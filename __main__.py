import logging
import os
import tkinter as tk
import pandas as pd
import openpyxl as xl
from tkinter import filedialog, messagebox
from convert_to_excel import extract_tags
from get_file_open import file_open
from openpyxl.styles.borders import Border, Side

class Disp:
    try:
        def __init__(self, root):
                self.root = root
                self.root.title('GTM CONVERT APP')
                self.selected_file_path = ""

                # ファイル選択ボタン
                self.file_frame = tk.Frame(self.root)
                self.file_frame.pack(pady=10)
                tk.Label(self.file_frame, text="ファイル選択:").grid(row=0, column=0, padx=5)
                self.file_path = tk.StringVar()  # ファイルパスを格納するためのStringVarを定義
                self.file_entry = tk.Entry(self.file_frame, width=50, textvariable=self.file_path, state='readonly')  # textvariableを使用してファイルパスをエントリに関連付ける
                self.file_entry.grid(row=0, column=1, padx=5)
                self.select_file_button = tk.Button(self.file_frame, text="選択", command=self.select_file)
                self.select_file_button.grid(row=0, column=2, padx=5)

                # 実行ボタン
                self.execute_button = tk.Button(self.root, text="実行", command=self.execute)
                self.execute_button.pack(pady=10)

        # ファイルの選択フィールド
        def select_file(self):
            file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
            if file_path:
                self.selected_file_path = file_path
                filename = os.path.basename(file_path)
                self.file_path.set(filename)  # ファイル名をStringVarに設定する
        # 実行ボタンがクリックされたときの処理
        def execute(self):
            
            file_data = None
            try:
                if not self.selected_file_path:
                    messagebox.showwarning("エラー","ファイルが選択されていません。")
                    return
                
                file_data = file_open(self.selected_file_path)
                if not file_data:
                    messagebox.showerror("エラー","ファイルの読み込みに失敗しました。")
                    return

                df = extract_tags(file_data)
                if df is None or df.empty:
                    messagebox.showerror("エラー","データフレームの生成に失敗しました。")
                    return

                # 保存ダイアログを表示してファイルを保存
                save_path = filedialog.asksaveasfilename(confirmoverwrite=True, defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                if save_path:
                    # ExcelWriterオブジェクトを作成
                    excel_writer = pd.ExcelWriter(save_path, engine='xlsxwriter')

                    # DataFrameをExcelファイルに書き込む
                    df.to_excel(excel_writer, index=False, sheet_name='タグ', header=True, engine='xlsxwriter')

                    # シートオブジェクトを取得
                    worksheet = excel_writer.sheets['タグ']

                    # DataFrameの行数を取得
                    num_rows = len(df)

                    # 値が入っている行のインデックスを格納するリスト
                    non_empty_rows = []

                    # 値が入っている行のインデックスを取得
                    for row_idx in range(1, num_rows):
                        if not df.iloc[row_idx].isnull().all():
                            non_empty_rows.append(row_idx)

                    # 値が入っている行の高さを設定
                    for row_idx in non_empty_rows:
                        worksheet.set_row(row_idx, 175)

                    # カラムの幅を設定
                    cell_format = excel_writer.book.add_format({'text_wrap': True, 'valign': 'top'})
                    worksheet.set_column('A:A', 4.5, cell_format)  # A列の幅を設定
                    worksheet.set_column('B:B', 17.5, cell_format)  # B列の幅を設定
                    worksheet.set_column('C:C', 74.5, cell_format)  # C列の幅を設定
                    worksheet.set_column('D:D', 24.5, cell_format)  # D列の幅を設定
                    worksheet.set_column('E:E', 54.5, cell_format)  # E列の幅を設定

                    # 罫線の設定
                    side = Side(style='thin', color='000000')

                    border = Border(top=side, bottom=side, left=side, right=side)

                    # 罫線を引く
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if worksheet.iter_rows()[cell.coordinate].value:
                                worksheet.iter_rows()[cell.coordinate].border = border

                    # ExcelWriterオブジェクトを保存して閉じる
                    excel_writer.close()

                    messagebox.showinfo("Success", "ファイルが保存されました。")
                else:
                    messagebox.showwarning("Warning", "ファイルの保存がキャンセルされました。")

            except Exception as e:
                messagebox.showerror("エラー",f"予期しないエラーが発生しました: {e}")
                logging.error(f"関数displayの呼び出しに失敗しました:{e}")
                
    except Exception as e:
        logging.error(f"関数displayの呼び出しに失敗しました:{e}")
    
if __name__ == "__main__":
    root = tk.Tk()
    app = Disp(root)
    root.mainloop()
